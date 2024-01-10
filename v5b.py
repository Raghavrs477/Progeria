import difflib
import tkinter as tk
from tkinter import messagebox
import openpyxl
import cv2
import dlib

# Define the symptoms
symptoms = [
    "Fever",
    "Cough",
    "Headache",
    "Fatigue",
    "Sore throat",
    "Shortness of breath",
    "Muscle or body aches",
    "Loss of taste or smell",
    "Nausea or vomiting",
    "Diarrhea"
]

# Function to evaluate symptoms and determine the condition
def evaluate_symptoms():
    selected_symptoms = [symptom.get() for symptom in symptom_vars.values()]
    selected_symptoms = [s for s in selected_symptoms if s != '']
    
    match_ratio = difflib.SequenceMatcher(None, selected_symptoms, symptoms).ratio()
    
    if match_ratio > 0.5:
        messagebox.showinfo("Evaluation Result", "You may have a condition related to the symptoms.")
    else:
        messagebox.showinfo("Evaluation Result", "You are fine.")
    
    # Add user information and symptoms to the Excel file
    add_user_record(entry_name.get(), entry_email.get(), selected_symptoms, match_ratio)

# Function to add user record to the Excel file
def add_user_record(name, email, symptoms, percentage):
    wb = openpyxl.load_workbook('user_records.xlsx')

    # Select the active sheet
    sheet = wb.active

    # Get the last row index
    last_row = sheet.max_row

    # Add user information to the Excel file
    sheet.cell(row=last_row+1, column=1, value=name)
    sheet.cell(row=last_row+1, column=2, value=email)
    sheet.cell(row=last_row+1, column=3, value=", ".join(symptoms))
    sheet.cell(row=last_row+1, column=4, value=percentage)

    # Save the Excel file
    wb.save('user_records.xlsx')

# Function to perform face recognition
def perform_face_recognition():
    detector = dlib.get_frontal_face_detector()

    # Load the preprocessed image
    preprocessed_image = cv2.imread('preprocessed_image.jpg')

    # Start the camera capture
    capture = cv2.VideoCapture(0)

    while True:
        ret, frame = capture.read()

        if ret:
            gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
            faces = detector(gray)

            for face in faces:
                (x, y, w, h) = (face.left(), face.top(), face.width(), face.height())
                cv2.rectangle(frame, (x, y), (x+w, y+h), (0, 255, 0), 2)

            cv2.imshow('Face Recognition', frame)

            if cv2.waitKey(1) == ord('q'):
                break

    capture.release()
    cv2.destroyAllWindows()

    # Perform face recognition by comparing the captured frame with the preprocessed image
    # Add your face recognition algorithm here

    # Prompt the result to the user
    messagebox.showinfo("Face Recognition Result", "You are suffering from Progeria.")

# Create the GUI window
window = tk.Tk()
window.title("Symptom Evaluator")

# Create and pack the registration labels and text entry fields
label_name = tk.Label(window, text="Enter your name:", font=("Arial", 14))
label
