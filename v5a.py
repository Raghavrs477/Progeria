import difflib
import tkinter as tk
from tkinter import messagebox
import openpyxl

# Define the symptoms
symptoms = [
    "Fever",
    "Cough",
    "Headache",
    "Fatigue",
    "Sore throat",
    "Shortness of breath",
    "Muscle or body aches",
    "Loss of taste or smell",
    "Nausea or vomiting",
    "Diarrhea"
]

# Function to evaluate symptoms and determine the condition
def evaluate_symptoms():
    selected_symptoms = [symptom.get() for symptom in symptom_vars.values()]
    selected_symptoms = [s for s in selected_symptoms if s != '']
    
    match_ratio = difflib.SequenceMatcher(None, selected_symptoms, symptoms).ratio()
    
    if match_ratio > 0.5:
        messagebox.showinfo("Evaluation Result", "You may have a condition related to the symptoms.")
    else:
        messagebox.showinfo("Evaluation Result", "You are fine.")
    
    # Add user information and symptoms to the Excel file
    add_user_record(entry_name.get(), entry_email.get(), selected_symptoms, match_ratio)

# Function to add user record to the Excel file
def add_user_record(name, email, symptoms, percentage):
    wb = openpyxl.load_workbook('user_records.xlsx')

    # Select the active sheet
    sheet = wb.active

    # Get the last row index
    last_row = sheet.max_row

    # Add user information to the Excel file
    sheet.cell(row=last_row+1, column=1, value=name)
    sheet.cell(row=last_row+1, column=2, value=email)
    sheet.cell(row=last_row+1, column=3, value=", ".join(symptoms))
    sheet.cell(row=last_row+1, column=4, value=percentage)

    # Save the Excel file
    wb.save('user_records.xlsx')

# Create the GUI window
window = tk.Tk()
window.title("Symptom Evaluator")

# Create and pack the registration labels and text entry fields
label_name = tk.Label(window, text="Enter your name:", font=("Arial", 14))
label_name.pack(pady=10)

entry_name = tk.Entry(window, width=30, font=("Arial", 12))
entry_name.pack()

label_email = tk.Label(window, text="Enter your email address:", font=("Arial", 14))
label_email.pack(pady=10)

entry_email = tk.Entry(window, width=30, font=("Arial", 12))
entry_email.pack()

# Create and pack the registration button
button_register = tk.Button(window, text="Register", command=lambda: register_symptoms(), font=("Arial", 14))
button_register.pack(pady=10)

# Create and pack the symptom selection label
label_symptoms = tk.Label(window, text="Select your symptoms:", font=("Arial", 14))
label_symptoms.pack(pady=10)

# Create a frame to hold the symptom selection checkboxes
symptoms_frame = tk.Frame(window)
symptoms_frame.pack()

# Create variables for symptom selection checkboxes
symptom_vars = {}
for symptom in symptoms:
    var = tk.StringVar()
    symptom_vars[symptom] = var
    checkbox = tk.Checkbutton(symptoms_frame, text=symptom, variable=var, onvalue=symptom, offvalue='',